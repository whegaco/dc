import streamlit as st
import pandas as pd
import re
import json
import hashlib
from pathlib import Path
from datetime import date
from openpyxl import load_workbook

st.set_page_config(page_title="IR Tracker System", layout="wide")

LOG_FILE = Path("01- Inspection Request - IR.xlsx")
USERS_FILE = Path("users.json")
DATA_SHEETS = ["STR", "ARCH", "ELEC", "MECH", "SURV"]
DATA_START_ROW = {"STR": 9, "ARCH": 9, "ELEC": 9, "MECH": 8, "SURV": 9}

CODE_MAP = {
    "A": "Approved",
    "B": "Approved with notes",
    "C": "Reject & Resubmit",
    "CC": "Not Resubmitted",
    "D": "Rejected Final",
    "UR": "Under Review",
    "S": "Superseded",
}

def hash_text(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()

def load_users():
    if not USERS_FILE.exists():
        st.error("users.json not found beside app.py")
        st.stop()
    with USERS_FILE.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_users(users):
    with USERS_FILE.open("w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def verify_login(username: str, password: str):
    users = load_users()
    user = users.get(username)
    if not user:
        return None
    return user if user.get("password_hash") == hash_text(password) else None

def norm_code(v):
    return str(v).strip().upper() if v is not None else ""

def extract_building(location):
    if not location:
        return ""
    text = str(location)
    m = re.search(r'Building\s*:?\s*([A-Za-z0-9()\-_/ ]+)', text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def extract_zone(location):
    if not location:
        return ""
    text = str(location)
    m = re.search(r'Zone\s*:?\s*([A-Za-z0-9\-_ ]+)', text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def delay_status(days):
    if pd.isna(days):
        return "Pending"
    if days <= 2:
        return "On Time"
    return "Delayed"

@st.cache_data
def load_all_data(file_mtime):
    wb = load_workbook(LOG_FILE, data_only=False)
    all_rows = []
    for sheet_name in DATA_SHEETS:
        ws = wb[sheet_name]
        start_row = DATA_START_ROW[sheet_name]
        for row in range(start_row, ws.max_row + 1):
            ref = ws.cell(row, 2).value
            if ref in (None, ""):
                continue
            code = norm_code(ws.cell(row, 11).value)
            rev = ws.cell(row, 3).value
            row_data = {
                "Sheet": sheet_name,
                "Excel Row": row,
                "No.": ws.cell(row, 1).value,
                "Reference": str(ref).strip(),
                "REV": pd.to_numeric(rev, errors="coerce"),
                "Element": ws.cell(row, 4).value,
                "Location": ws.cell(row, 5).value,
                "Description": ws.cell(row, 6).value,
                "Dwg": ws.cell(row, 7).value,
                "Report": ws.cell(row, 8).value,
                "Submission date": ws.cell(row, 9).value,
                "Received Date": ws.cell(row, 10).value,
                "Code Action": code,
                "Code Meaning": CODE_MAP.get(code, ""),
            }
            row_data["Building"] = extract_building(row_data["Location"])
            row_data["Zone"] = extract_zone(row_data["Location"])
            all_rows.append(row_data)

    df = pd.DataFrame(all_rows)
    if df.empty:
        return df, df, df

    df["Submission date"] = pd.to_datetime(df["Submission date"], errors="coerce")
    df["Received Date"] = pd.to_datetime(df["Received Date"], errors="coerce")
    df["Delay Days"] = (df["Received Date"] - df["Submission date"]).dt.days
    df["Delay Status"] = df["Delay Days"].apply(delay_status)

    raw_df = df.copy()
    latest_df = (
        df.sort_values(["Sheet", "Reference", "REV", "Excel Row"], na_position="last")
          .groupby(["Sheet", "Reference"], as_index=False)
          .tail(1)
          .reset_index(drop=True)
    )

    active_df = latest_df[
        latest_df["Code Action"].astype(str).str.strip().ne("") &
        latest_df["Code Action"].astype(str).str.upper().ne("S")
    ].copy()

    return raw_df, latest_df, active_df

def next_reference(df, sheet_name):
    prefix = f"WIR-{sheet_name}-"
    nums = []
    for val in df[df["Sheet"] == sheet_name]["Reference"].dropna().astype(str):
        v = val.upper().strip()
        if v.startswith(prefix):
            try:
                nums.append(int(v.split("-")[-1]))
            except Exception:
                pass
    nxt = max(nums) + 1 if nums else 1
    return f"{prefix}{nxt:03d}"

def append_to_sheet(record):
    wb = load_workbook(LOG_FILE)
    ws = wb[record["Sheet"]]
    next_row = ws.max_row + 1
    data_start = DATA_START_ROW[record["Sheet"]]

    ws.cell(next_row, 1).value = next_row - data_start + 1
    ws.cell(next_row, 2).value = record["Reference"]
    ws.cell(next_row, 3).value = record["REV"]
    ws.cell(next_row, 4).value = record["Element"]
    ws.cell(next_row, 5).value = record["Location"]
    ws.cell(next_row, 6).value = record["Description"]
    ws.cell(next_row, 7).value = record["Dwg"]
    ws.cell(next_row, 8).value = record["Report"]
    ws.cell(next_row, 9).value = record["Submission date"]
    ws.cell(next_row, 10).value = record["Received Date"]
    ws.cell(next_row, 11).value = record["Code Action"]
    wb.save(LOG_FILE)

def update_sheet_row(sheet_name, excel_row, new_rev, new_code, new_received_date, new_description):
    wb = load_workbook(LOG_FILE)
    ws = wb[sheet_name]
    ws.cell(excel_row, 3).value = int(new_rev)
    ws.cell(excel_row, 10).value = new_received_date
    ws.cell(excel_row, 11).value = new_code
    if new_description is not None:
        ws.cell(excel_row, 6).value = new_description
    wb.save(LOG_FILE)

def login_page():
    st.title("IR Tracker Login")
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Login")
    if submit:
        user = verify_login(username.strip(), password)
        if user:
            st.session_state["logged_in"] = True
            st.session_state["username"] = username.strip()
            st.session_state["role"] = user.get("role", "viewer")
            st.rerun()
        else:
            st.error("اسم المستخدم أو كلمة السر غلط")

def logout_button():
    if st.sidebar.button("Logout"):
        for key in ["logged_in", "username", "role"]:
            st.session_state.pop(key, None)
        st.rerun()

def dashboard(raw_df, latest_df, active_df):
    st.title("IR Tracker Dashboard")

    c0, c1, c2, c3, c4, c5 = st.columns(6)
    c0.metric("Total Active IR", len(active_df))
    c1.metric("B", int((active_df["Code Action"] == "B").sum()))
    c2.metric("C / CC", int((active_df["Code Action"].isin(["C", "CC"])).sum()))
    c3.metric("D", int((active_df["Code Action"] == "D").sum()))
    c4.metric("UR", int((active_df["Code Action"] == "UR").sum()))
    c5.metric("Delayed", int((active_df["Delay Status"] == "Delayed").sum()))

    with st.expander("Count Logic"):
        st.write(f"Raw rows in sheets: {len(raw_df)}")
        st.write(f"Latest revision only: {len(latest_df)}")
        st.write(f"Active IR (latest + coded + excluding S): {len(active_df)}")

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("IR Count by Sheet")
        sheet_counts = active_df["Sheet"].value_counts().rename_axis("Sheet").reset_index(name="Count")
        st.bar_chart(sheet_counts.set_index("Sheet"))
    with c2:
        st.subheader("Code Action Summary")
        code_counts = active_df["Code Action"].value_counts().rename_axis("Code").reset_index(name="Count")
        st.dataframe(code_counts, use_container_width=True, hide_index=True)

    c3, c4 = st.columns(2)
    with c3:
        st.subheader("Delay Status Summary")
        delay_counts = active_df["Delay Status"].value_counts().rename_axis("Delay Status").reset_index(name="Count")
        st.dataframe(delay_counts, use_container_width=True, hide_index=True)
    with c4:
        st.subheader("Delayed IRs")
        delayed = active_df[active_df["Delay Status"] == "Delayed"].copy()
        cols = ["Sheet", "Reference", "Building", "Delay Days", "Code Action"]
        st.dataframe(delayed[cols].sort_values("Delay Days", ascending=False).head(15), use_container_width=True, hide_index=True)

def add_ir_page(active_df):
    st.title("Add Inspection Request")
    with st.form("add_ir"):
        c1, c2, c3 = st.columns(3)
        with c1:
            sheet_name = st.selectbox("Sheet / Discipline", DATA_SHEETS)
            auto_ref = st.checkbox("Generate Reference automatically", value=True)
            rev = st.number_input("REV", min_value=0, value=0, step=1)
            element = st.text_input("Element", value="")
        with c2:
            zone = st.text_input("Zone", value="04")
            building = st.text_input("Building", value="VF-01")
            submission_date = st.date_input("Submission date", value=date.today())
            received_date = st.date_input("Received Date", value=date.today())
        with c3:
            code_action = st.selectbox("Code Action", ["A", "B", "C", "CC", "D", "UR"])
            dwg = st.text_input("Dwg", value="")
            report = st.text_input("Report", value="")
        description = st.text_area("Description", value="")
        if auto_ref:
            ref = next_reference(active_df, sheet_name)
            st.info(f"Generated Reference: {ref}")
        else:
            ref = st.text_input("Reference", value="")
        submit = st.form_submit_button("Add to IR Log")

    if submit:
        if not ref.strip():
            st.error("Reference مطلوب")
            return
        exists = active_df["Reference"].dropna().astype(str).str.upper().eq(ref.upper()).any()
        if exists:
            st.error("الـ Reference ده موجود قبل كدا")
            return
        record = {
            "Sheet": sheet_name,
            "Reference": ref.strip(),
            "REV": int(rev),
            "Element": element.strip(),
            "Location": f"Zone: {zone}\nBuilding: {building}",
            "Description": description.strip(),
            "Dwg": dwg if dwg != "" else None,
            "Report": report if report != "" else None,
            "Submission date": submission_date,
            "Received Date": received_date,
            "Code Action": code_action,
        }
        append_to_sheet(record)
        st.success(f"تمت إضافة {ref} في شيت {sheet_name}")
        st.cache_data.clear()
        st.rerun()

def update_ir_page(active_df):
    st.title("Update IR")
    refs = active_df["Reference"].dropna().astype(str).sort_values().tolist()
    selected_ref = st.selectbox("Select IR Reference", refs)
    row = active_df[active_df["Reference"] == selected_ref].iloc[0]

    st.caption(f"Sheet: {row['Sheet']} | Excel Row: {int(row['Excel Row'])}")

    c1, c2, c3 = st.columns(3)
    with c1:
        new_rev = st.number_input("REV", min_value=0, value=int(row["REV"]) if not pd.isna(row["REV"]) else 0, step=1)
    with c2:
        code_options = ["A", "B", "C", "CC", "D", "UR"]
        current_code = row["Code Action"] if row["Code Action"] in code_options else "UR"
        new_code = st.selectbox("Code Action", code_options, index=code_options.index(current_code))
    with c3:
        current_received = row["Received Date"].date() if pd.notna(row["Received Date"]) else date.today()
        new_received_date = st.date_input("Received Date", value=current_received)

    new_description = st.text_area("Description", value="" if pd.isna(row["Description"]) else str(row["Description"]))

    st.write("Current Data")
    current_cols = ["Reference", "REV", "Code Action", "Delay Days", "Delay Status", "Building", "Description"]
    st.dataframe(pd.DataFrame([row[current_cols]]), use_container_width=True, hide_index=True)

    if st.button("Save Update"):
        update_sheet_row(
            sheet_name=row["Sheet"],
            excel_row=int(row["Excel Row"]),
            new_rev=int(new_rev),
            new_code=new_code,
            new_received_date=new_received_date,
            new_description=new_description,
        )
        st.success("تم تحديث الـ IR")
        st.cache_data.clear()
        st.rerun()

def reminder_page(active_df):
    st.title("Reminder / Follow-up")
    st.caption("الحالات اللي محتاجة متابعة")

    pending_follow = active_df[
        (active_df["Delay Status"] == "Delayed") |
        (active_df["Code Action"].isin(["UR", "C", "CC"]))
    ].copy()

    c1, c2, c3 = st.columns(3)
    c1.metric("Need Follow-up", len(pending_follow))
    c2.metric("Under Review", int((pending_follow["Code Action"] == "UR").sum()))
    c3.metric("Re-submit Pending", int((pending_follow["Code Action"].isin(["C", "CC"])).sum()))

    show_cols = ["Sheet", "Reference", "REV", "Building", "Description", "Submission date", "Received Date", "Delay Days", "Delay Status", "Code Action"]
    st.dataframe(
        pending_follow[show_cols].sort_values(["Delay Status", "Delay Days"], ascending=[False, False]),
        use_container_width=True,
        hide_index=True,
    )

    csv = pending_follow[show_cols].to_csv(index=False).encode("utf-8-sig")
    st.download_button("Download reminder list CSV", data=csv, file_name="IR_followup_list.csv", mime="text/csv")

def building_report_page(active_df):
    st.title("Building Report")
    buildings = sorted([b for b in active_df["Building"].dropna().astype(str).unique().tolist() if b.strip()])
    if not buildings:
        st.warning("مفيش Buildings متقراية من اللوج")
        return
    selected = st.selectbox("Select Building", buildings)
    filtered = active_df[active_df["Building"].astype(str) == selected].copy()

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Total", len(filtered))
    c2.metric("B", int((filtered["Code Action"] == "B").sum()))
    c3.metric("C/CC", int((filtered["Code Action"].isin(["C", "CC"])).sum()))
    c4.metric("D", int((filtered["Code Action"] == "D").sum()))
    c5.metric("UR", int((filtered["Code Action"] == "UR").sum()))
    c6.metric("Delayed", int((filtered["Delay Status"] == "Delayed").sum()))

    show_cols = ["Sheet", "Reference", "REV", "Element", "Zone", "Building", "Description", "Submission date", "Received Date", "Delay Days", "Delay Status", "Code Action", "Code Meaning"]
    st.dataframe(filtered[show_cols], use_container_width=True, hide_index=True)

def code_report_page(active_df):
    st.title("Code Report")
    code_action = st.selectbox("Select Code", ["A", "B", "C", "CC", "D", "UR"])
    filtered = active_df[active_df["Code Action"] == code_action].copy()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total", len(filtered))
    c2.metric("Buildings", int(filtered["Building"].fillna("").astype(str).replace("", pd.NA).dropna().nunique()))
    c3.metric("Sheets", int(filtered["Sheet"].nunique()))
    avg_rev = round(pd.to_numeric(filtered["REV"], errors="coerce").dropna().mean(), 2) if not filtered.empty else 0
    c4.metric("Avg REV", avg_rev)

    show_cols = ["Sheet", "Reference", "REV", "Element", "Zone", "Building", "Description", "Submission date", "Received Date", "Delay Days", "Delay Status", "Code Action", "Code Meaning"]
    st.dataframe(filtered[show_cols], use_container_width=True, hide_index=True)

    c5, c6 = st.columns(2)
    with c5:
        sheet_counts = filtered["Sheet"].value_counts().rename_axis("Sheet").reset_index(name="Count")
        if not sheet_counts.empty:
            st.bar_chart(sheet_counts.set_index("Sheet"))
    with c6:
        bld = filtered["Building"].fillna("").astype(str)
        bld = bld[bld.str.strip() != ""].value_counts().rename_axis("Building").reset_index(name="Count").head(15)
        st.dataframe(bld, use_container_width=True, hide_index=True)

def search_page(active_df):
    st.title("Search IR")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        ref = st.text_input("Reference")
    with c2:
        building = st.text_input("Building")
    with c3:
        sheet_name = st.selectbox("Sheet", [""] + DATA_SHEETS)
    with c4:
        code_action = st.selectbox("Code", ["", "A", "B", "C", "CC", "D", "UR"])

    keyword = st.text_input("Keyword in description/location")
    result = active_df.copy()

    if ref:
        result = result[result["Reference"].fillna("").astype(str).str.contains(ref, case=False, na=False)]
    if building:
        result = result[result["Building"].fillna("").astype(str).str.contains(building, case=False, na=False)]
    if sheet_name:
        result = result[result["Sheet"] == sheet_name]
    if code_action:
        result = result[result["Code Action"] == code_action]
    if keyword:
        blob = result["Description"].fillna("").astype(str) + " " + result["Location"].fillna("").astype(str)
        result = result[blob.str.contains(keyword, case=False, na=False)]

    st.write(f"Results: {len(result)}")
    st.dataframe(result, use_container_width=True, hide_index=True)

def manage_users_page():
    st.title("Manage Viewer Users")
    users = load_users()

    rows = [{"Username": username, "Role": data.get("role", "viewer")} for username, data in users.items()]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with st.form("add_viewer_form"):
        st.subheader("Add Viewer")
        new_username = st.text_input("New viewer username")
        new_password = st.text_input("New viewer password", type="password")
        add_submit = st.form_submit_button("Add Viewer")
    if add_submit:
        new_username = new_username.strip()
        if not new_username or not new_password:
            st.error("لازم تكتب اسم المستخدم وكلمة السر")
            return
        if new_username in users:
            st.error("اسم المستخدم موجود قبل كدا")
            return
        users[new_username] = {"password_hash": hash_text(new_password), "role": "viewer"}
        save_users(users)
        st.success(f"تم إضافة المستخدم {new_username} كـ viewer")
        st.rerun()

    viewer_users = [u for u, d in users.items() if d.get("role") == "viewer"]
    if viewer_users:
        st.subheader("Delete Viewer")
        del_user = st.selectbox("Select viewer to delete", viewer_users)
        if st.button("Delete Viewer"):
            users.pop(del_user, None)
            save_users(users)
            st.success(f"تم حذف المستخدم {del_user}")
            st.rerun()

def legend():
    st.markdown("""
    **Code legend**
    - A = Approved
    - B = Approved with notes
    - C = Reject & Resubmit
    - CC = Not Resubmitted
    - D = Rejected Final
    - UR = Under Review
    - S = Superseded (excluded)
    """)

def main():
    if not LOG_FILE.exists():
        st.error("حط ملف 01- Inspection Request - IR.xlsx جنب app.py")
        st.stop()

    if not st.session_state.get("logged_in", False):
        login_page()
        st.stop()

    raw_df, latest_df, active_df = load_all_data(LOG_FILE.stat().st_mtime)
    role = st.session_state.get("role", "viewer")
    username = st.session_state.get("username", "")

    with st.sidebar:
        st.title("IR Tracker")
        st.caption(f"Logged in as: {username} ({role})")
        if role == "admin":
            page = st.radio("Pages", ["Dashboard", "Add IR", "Update IR", "Reminder", "Building Report", "Code Report", "Search", "Manage Users"])
        else:
            page = st.radio("Pages", ["Dashboard", "Reminder", "Building Report", "Code Report", "Search"])
        legend()
        st.caption(f"Active IR loaded: {len(active_df)}")
        logout_button()

    if page == "Dashboard":
        dashboard(raw_df, latest_df, active_df)
    elif page == "Add IR" and role == "admin":
        add_ir_page(active_df)
    elif page == "Update IR" and role == "admin":
        update_ir_page(active_df)
    elif page == "Reminder":
        reminder_page(active_df)
    elif page == "Building Report":
        building_report_page(active_df)
    elif page == "Code Report":
        code_report_page(active_df)
    elif page == "Manage Users" and role == "admin":
        manage_users_page()
    else:
        search_page(active_df)

if __name__ == "__main__":
    main()
